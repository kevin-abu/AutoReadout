
import openpyxl
wb1 = openpyxl.load_workbook("XMEGA_PRODSIG_FUSES_CONFIG.xlsx")
#wb1 = openpyxl.load_workbook('XMEGA_PRODSIG_FUSES_CONFIG.xlsx')
ws1_prod_sig = wb1['XMEGA_PRODSIG']
ws1_fuses = wb1['XMEGA_FUSES']
wb2 = openpyxl.load_workbook("XMEGA_READOUT_TEMPLATE.xlsx")
#wb2 = openpyxl.load_workbook('XMEGA_READOUT_TEMPLATE.xlsx')
ws2 = wb2['Sheet1']

def exceltohtml(workbook_save,new_flabel):
    from xlsx2html import xlsx2html
    xlsx2html(workbook_save, new_flabel + "_Readout-Summary.html")

    

#######################-------SETUP-------#######################
def XMEGA_A_SETUP(workbook_save):
    
    for i in range (2, 56): 
        for j in range (1, 2): 
            # reading cell value from source excel file 
            c_prod_sig = ws1_prod_sig.cell(row = i, column = j)
            print (c_prod_sig.value)
            ws2.cell(row = i+18, column = j+2).value = c_prod_sig.value # write production signatures
            ws2.cell(row = i+70, column = j+3).value = None
    for i in range (2, 8): 
        for j in range (1,2):
            c_fuses = ws1_fuses.cell(row = i, column = j)
            ws2.cell(row = i+11, column = j+2).value = c_fuses.value # write fuses
    wb2.save("XMEGA_READOUT_TEMPLATE_A.xlsx")
    #wb2.save(workbook_save)
    

def XMEGA_B_SETUP(workbook_save):
    #FOR XMEGA-B / XMEGA - C#
    for i in range (2, 56): 
        for j in range (3,4): 
            # reading cell value from source excel file 
            c_prod_sig = ws1_prod_sig.cell(row = i, column = j)
            print (c_prod_sig.value)
            ws2.cell(row = i+18, column = j).value = c_prod_sig.value # write production signatures
            ws2.cell(row = i+66, column = j+1).value = None
    for i in range (2, 8): 
        for j in range (3,4):
            c_fuses = ws1_fuses.cell(row = i, column = j)
            ws2.cell(row = i+11, column = j).value = c_fuses.value # write fuses
    wb2.save("XMEGA_READOUT_TEMPLATE_B.xlsx")
    #wb2.save(workbook_save)
    
    ##
def XMEGA_C_SETUP(workbook_save):
    #FOR XMEGA-B / XMEGA - C#
    for i in range (2, 56): 
        for j in range (3,4): 
            # reading cell value from source excel file 
            c_prod_sig = ws1_prod_sig.cell(row = i, column = j)
            print (c_prod_sig.value)
            ws2.cell(row = i+18, column = j).value = c_prod_sig.value # write production signatures
            ws2.cell(row = i+66, column = j+1).value = None
    for i in range (2, 8): 
        for j in range (3,4):
            c_fuses = ws1_fuses.cell(row = i, column = j)
            ws2.cell(row = i+11, column = j).value = c_fuses.value # write fuses
    wb2.save("XMEGA_READOUT_TEMPLATE_C.xlsx")
    #wb2.save(workbook_save)
    
    ##
def XMEGA_D_SETUP(workbook_save):
    #FOR XMEGA D#
    for i in range (2, 56): 
        for j in range (7,8): 
            # reading cell value from source excel file 
            c_prod_sig = ws1_prod_sig.cell(row = i, column = j)
            print (c_prod_sig.value)
            ws2.cell(row = i+18, column = j-4).value = c_prod_sig.value # write production signatures
            ws2.cell(row = i+66, column = j-3).value = None
    for i in range (2, 8): 
        for j in range (7,8):
            c_fuses = ws1_fuses.cell(row = i, column = j)
            ws2.cell(row = i+11, column = j-4).value = c_fuses.value # write fuses
    wb2.save("XMEGA_READOUT_TEMPLATE_D.xlsx")
    #wb2.save(workbook_save)

def XMEGA_E_SETUP(workbook_save):
#FOR XMEGA E#
    for i in range (2, 56): 
        for j in range (9, 10): 
            # reading cell value from source excel file 
            c_prod_sig = ws1_prod_sig.cell(row = i, column = j)
            print (c_prod_sig.value)
            ws2.cell(row = i+18, column = j-6).value = c_prod_sig.value # write production signatures
    for i in range (2, 8): 
        for j in range (9, 10):
            c_fuses = ws1_fuses.cell(row = i, column = j)
            ws2.cell(row = i+11, column = j-6).value = c_fuses.value # write fuses

    wb2.save("READOUT_AUTOMATION\\XMEGA_READOUT_TEMPLATE_E.xlsx")
    #wb2.save(workbook_save)
    ##
print ('Ok')
#######################-------SETUP-------#######################

#######################-------FUSES-------#######################

def XMEGA_FUSES(new_flabel,workbook_save):#FOR ALL XMEGA VERSIONS
    #CHANGE TO OUTPUT FILE#
    #xmega_fuse=open(new_flabel  + "_fuses.hex",'r').read()
    wb2 = openpyxl.load_workbook(workbook_save)
    ws2 = wb2['Sheet1']
    xmega_fuse=open("C:\\Users\\a50291\\Documents\\READOUT_AUTOMATION\\2_2020-000006_SN1_1.6_ATXMEGA32E5_fuses.hex",'r').read()# FUSES FILE
    #xmega_fuse=open(new_flabel  + "_fuses.hex",'r').read()
    x_fuse= []
    for i in range (9,23,2): 
        x_fuse.append(xmega_fuse[i:i+2])    
    print ("Writing fuses.")        
    fuse_value = x_fuse[0:3] + x_fuse[4:7] 
    for i in range (1,7): 
        ws2.cell(row = i+12, column = 4).value = "0x" + fuse_value[i-1]
    wb2.save("C:\\Users\\a50291\\Documents\\READOUT_AUTOMATION\\XMEGA_READOUT_TEMPLATE_A.xlsx")
    #wb2.save(workbook_save)
    
#######################-------FUSES-------#######################

    
#######################-------PROD SIG-------#######################
    
def XMEGA_E_PROD_SIG(new_flabel,workbook_save):#XMEGA E
    #xmega_prod_signature=open(new_flabel  + "_prodsig.hex",'r').read()
    wb2 = openpyxl.load_workbook(workbook_save)
    ws2 = wb2['Sheet1']
    xmega_prod_signature=open("C:\\Users\\a50291\\Documents\\READOUT_AUTOMATION\\4_2020-000006_SN1_1.6_ATXMEGA32E5_prodsig.hex",'r').read()# PRODSIG FILE
    x_prod_sig= []
    for i in range (9,41,2): #Loop1
        x_prod_sig.append(xmega_prod_signature[i:i+2])    
    for i in range (53,85,2): #Loop2
        x_prod_sig.append(xmega_prod_signature[i:i+2])
    for i in range (97,128,2): #Loop3
        x_prod_sig.append(xmega_prod_signature[i:i+2])
    for i in range (141,153,2): #Loop4
        x_prod_sig.append(xmega_prod_signature[i:i+2])
    print (x_prod_sig)
    print ("Writing production signatures.")
      
    for i in range (1,55): 
        ws2.cell(row = i+19, column = 4).value = "0x" + x_prod_sig[i-1]
        ws2.cell(row = i, column = 4).alignment = Alignment(horizontal='center')
    wb2.save("C:\\Users\\a50291\\Documents\\READOUT_AUTOMATION\\XMEGA_READOUT_TEMPLATE_E.xlsx")
    #wb2.save(workbook_save)

    exceltohtml(workbook_save,new_flabel)
    
def XMEGA_BCD_PROD_SIG(new_flabel,workbook_save):#XMEGA B-C-D
    #xmega_prod_signature=open(new_flabel  + "_prodsig.hex",'r').read()
    wb2 = openpyxl.load_workbook(workbook_save)
    ws2 = wb2['Sheet1']
    xmega_prod_signature=open("C:\\Users\\a50291\\Documents\\READOUT_AUTOMATION\\4_2020-000006_SN1_1.6_ATXMEGA32E5_prodsig.hex",'r').read()# FUSES FILE
    x_prod_sig= []
    for i in range (9,41,2): #Loop1
        x_prod_sig.append(xmega_prod_signature[i:i+2])    
    for i in range (53,85,2): #Loop2
        x_prod_sig.append(xmega_prod_signature[i:i+2])
    for i in range (97,128,2): #Loop3
        x_prod_sig.append(xmega_prod_signature[i:i+2])
    print (x_prod_sig)
      
    for i in range (1,49): 
        ws2.cell(row = i+19, column = 4).value = "0x" + x_prod_sig[i-1]
        ws2.cell(row = i, column = 4).alignment = Alignment(horizontal='center')
    wb2.save("C:\\Users\\a50291\\Documents\\READOUT_AUTOMATION\\XMEGA_READOUT_TEMPLATE_A.xlsx")
    #wb2.save(workbook_save)

    exceltohtml(workbook_save,new_flabel)
    
def XMEGA_A_PROD_SIG(new_flabel,workbook_save):#XMEGA A
    #xmega_prod_signature=open(new_flabel  + "_prodsig.hex",'r').read()
    wb2 = openpyxl.load_workbook(workbook_save)
    ws2 = wb2['Sheet1']
    xmega_prod_signature=open("C:\\Users\\a50291\\Documents\\READOUT_AUTOMATION\\4_2020-000006_SN1_1.6_ATXMEGA32E5_prodsig.hex",'r').read()# FUSES FILE
    x_prod_sig= []
    for i in range (9,41,2): #Loop1
        x_prod_sig.append(xmega_prod_signature[i:i+2])    
    for i in range (53,85,2): #Loop2
        x_prod_sig.append(xmega_prod_signature[i:i+2])
    for i in range (97,128,2): #Loop3
        x_prod_sig.append(xmega_prod_signature[i:i+2])
    for i in range (141,149,2): #Loop4
        x_prod_sig.append(xmega_prod_signature[i:i+2])
    print (x_prod_sig)
      
    for i in range (1,53): 
        ws2.cell(row = i+19, column = 4).value = "0x" + x_prod_sig[i-1]
        ws2.cell(row = i, column = 4).alignment = Alignment(horizontal='center')
    wb2.save("XMEGA_READOUT_TEMPLATE_A.xlsx")
    #wb2.save(workbook_save)

    exceltohtml(workbook_save,new_flabel)



def xmega_version(dev_name,new_flabel,workbook_save):
    if dev_name[9] == ('A' or 'a'):
        print (dev_name[9])
        XMEGA_A_SETUP(workbook_save)
        XMEGA_FUSES(new_flabel,workbook_save)
        XMEGA_A_PROD_SIG(new_flabel,workbook_save)
        
        
    elif dev_name[9] == ('B' or 'b'):
        print (dev_name[9])
        XMEGA_B_SETUP(workbook_save)
        XMEGA_FUSES(new_flabel,workbook_save)
        XMEGA_BCD_PROD_SIG(new_flabel,workbook_save)
        
    elif dev_name[9] == ('C' or 'c'):
        print (dev_name[9])
        XMEGA_C_SETUP(workbook_save)
        XMEGA_FUSES(new_flabel,workbook_save)
        XMEGA_BCD_PROD_SIG(new_flabel,workbook_save)
        
    elif dev_name[9] == ('D' or 'd'):
        print (dev_name[9])
        XMEGA_D_SETUP(workbook_save)
        XMEGA_FUSES(new_flabel,workbook_save)
        XMEGA_BCD_PROD_SIG(new_flabel,workbook_save)
        
    elif dev_name[9] == ('E' or 'e'):
        print (dev_name[9])
        XMEGA_E_SETUP(workbook_save)
        XMEGA_FUSES(new_flabel,workbook_save)
        XMEGA_E_PROD_SIG(new_flabel,workbook_save)
        
    else:
        if dev_name[10] == ('A' or 'a'):
            print (dev_name[10])
            XMEGA_A_SETUP(workbook_save)
            XMEGA_FUSES(new_flabel,workbook_save)
            XMEGA_A_PROD_SIG(new_flabel,workbook_save)

        elif dev_name[10] == ('B' or 'b'):
            print (dev_name[10])
            XMEGA_B_SETUP(workbook_save)
            XMEGA_FUSES(new_flabel,workbook_save)
            XMEGA_BCD_PROD_SIG(new_flabel,workbook_save)
            
        elif dev_name[10] == ('C' or 'c'):
            print (dev_name[10])
            XMEGA_C_SETUP(workbook_save)
            XMEGA_FUSES(new_flabel,workbook_save)
            XMEGA_BCD_PROD_SIG(new_flabel,workbook_save)
            
        elif dev_name[10] == ('D' or 'd'):
            print (dev_name[10])
            XMEGA_D_SETUP(workbook_save)
            XMEGA_FUSES(new_flabel,workbook_save)
            XMEGA_BCD_PROD_SIG(new_flabel,workbook_save)
            
        elif dev_name[10] == ('E' or 'e'):
            print (dev_name[10])
            XMEGA_E_SETUP(workbook_save)
            XMEGA_FUSES(new_flabel,workbook_save)
            XMEGA_E_PROD_SIG(new_flabel,workbook_save)
            
            
